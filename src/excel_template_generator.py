#!/usr/bin/env python3
"""
Excel Template Generator for Geothermal Digital Twin AI

Creates comprehensive Excel templates for project governance, data tracking,
and system monitoring. Generates professional templates with multiple sheets
for different aspects of the geothermal digital twin system.

Features:
1. Project Overview and Dashboard
2. Literature Corpus Tracking  
3. Image Corpus Management (as requested)
4. 3D Models and Data Inventory
5. Engineering Analysis Results
6. QA Performance Metrics
7. Evaluation Results Tracking
8. Risk Assessment Matrix
9. Development Timeline and Milestones
10. Configuration Management

Dependencies:
- openpyxl: Excel file creation and formatting
- pandas: Data structure handling
- datetime: Date and time utilities

Usage:
    python src/excel_template_generator.py [--output FILE] [--update-existing]
"""

import os
import sys
import json
import logging
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Any

# Excel processing
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, NamedStyle
)
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Data handling
import pandas as pd
import numpy as np

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Constants
EXCEL_DIR = Path("excel")
TEMPLATE_FILE = EXCEL_DIR / "geothermal_AI_template_v2.xlsx"
CACHE_DIR = Path("digital_twin/cache")


class GeothermalExcelGenerator:
    """Professional Excel template generator for geothermal digital twin governance."""
    
    def __init__(self):
        """Initialize the Excel generator."""
        self.setup_directories()
        self.workbook = Workbook()
        
        # Color scheme
        self.colors = {
            'primary': 'FF2E4BC4',      # Blue
            'secondary': 'FF28A745',    # Green  
            'accent': 'FFFD7E14',       # Orange
            'warning': 'FFFFC107',      # Yellow
            'danger': 'FFDC3545',       # Red
            'light': 'FFF8F9FA',        # Light gray
            'dark': 'FF343A40',         # Dark gray
            'header': 'FF17A2B8',       # Teal
            'success': 'FF198754'       # Dark green
        }
        
        self.setup_styles()
        
        logger.info("Excel template generator initialized")
    
    def setup_directories(self):
        """Create required directories."""
        EXCEL_DIR.mkdir(parents=True, exist_ok=True)
    
    def setup_styles(self):
        """Define Excel styles for consistent formatting."""
        # Header style
        header_style = NamedStyle(name="header")
        header_style.font = Font(bold=True, color="FFFFFF", size=12)
        header_style.fill = PatternFill(start_color=self.colors['header'], 
                                       end_color=self.colors['header'], 
                                       fill_type="solid")
        header_style.alignment = Alignment(horizontal="center", vertical="center")
        header_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Data style
        data_style = NamedStyle(name="data")
        data_style.font = Font(size=10)
        data_style.alignment = Alignment(horizontal="left", vertical="center")
        data_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Number style
        number_style = NamedStyle(name="number")
        number_style.font = Font(size=10)
        number_style.alignment = Alignment(horizontal="right", vertical="center")
        number_style.number_format = '#,##0.00'
        number_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Status styles
        status_good = NamedStyle(name="status_good")
        status_good.font = Font(bold=True, color="FFFFFF")
        status_good.fill = PatternFill(start_color=self.colors['success'], 
                                      end_color=self.colors['success'], 
                                      fill_type="solid")
        status_good.alignment = Alignment(horizontal="center", vertical="center")
        
        status_warning = NamedStyle(name="status_warning")
        status_warning.font = Font(bold=True, color="000000")
        status_warning.fill = PatternFill(start_color=self.colors['warning'], 
                                         end_color=self.colors['warning'], 
                                         fill_type="solid")
        status_warning.alignment = Alignment(horizontal="center", vertical="center")
        
        status_danger = NamedStyle(name="status_danger")
        status_danger.font = Font(bold=True, color="FFFFFF")
        status_danger.fill = PatternFill(start_color=self.colors['danger'], 
                                        end_color=self.colors['danger'], 
                                        fill_type="solid")
        status_danger.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add styles to workbook
        for style in [header_style, data_style, number_style, status_good, status_warning, status_danger]:
            if style.name not in self.workbook.named_styles:
                self.workbook.add_named_style(style)
    
    def create_project_overview_sheet(self):
        """Create project overview and dashboard sheet."""
        logger.info("Creating Project Overview sheet...")
        
        # Remove default sheet and create new one
        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])
        
        ws = self.workbook.create_sheet("Project_Overview", 0)
        
        # Title
        ws['A1'] = "Geothermal Digital Twin AI - Project Dashboard"
        ws['A1'].font = Font(bold=True, size=16, color=self.colors['primary'])
        ws.merge_cells('A1:H1')
        
        # Project Information
        ws['A3'] = "Project Information"
        ws['A3'].style = "header"
        ws.merge_cells('A3:B3')
        
        project_info = [
            ("Project Name", "Semurup Geothermal Digital Twin"),
            ("Location", "Semurup, Jambi, Sumatra, Indonesia"),
            ("Project Phase", "Development"),
            ("Last Updated", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("System Version", "v2.0"),
            ("Status", "Active")
        ]
        
        for i, (key, value) in enumerate(project_info, 4):
            ws[f'A{i}'] = key
            ws[f'B{i}'] = value
            ws[f'A{i}'].style = "data"
            ws[f'B{i}'].style = "data"
        
        # System Statistics
        ws['D3'] = "System Statistics"
        ws['D3'].style = "header"
        ws.merge_cells('D3:E3')
        
        stats = [
            ("Literature Documents", "=Literature_Corpus!$B$2"),
            ("Image Files", "=Image_Corpus!$B$2"),
            ("3D Models", "=Model_Data!$B$2"),
            ("Engineering Analyses", "=Engineering_Results!$B$2"),
            ("Total Questions Tested", "=QA_Performance!$B$2"),
            ("System Health Score", "=Evaluation_Results!$B$2")
        ]
        
        for i, (key, formula) in enumerate(stats, 4):
            ws[f'D{i}'] = key
            ws[f'E{i}'] = formula
            ws[f'D{i}'].style = "data"
            ws[f'E{i}'].style = "number"
        
        # Key Performance Indicators
        ws['G3'] = "Key Performance Indicators"
        ws['G3'].style = "header"
        ws.merge_cells('G3:H3')
        
        kpis = [
            ("Geothermal Capacity (MW)", "=Engineering_Results!$D$5"),
            ("Development Status", "=Engineering_Results!$D$6"),
            ("QA System Accuracy", "=QA_Performance!$D$5"),
            ("Response Time (avg)", "=QA_Performance!$D$6"),
            ("System Uptime", "=Evaluation_Results!$D$5"),
            ("Data Coverage", "=Model_Data!$D$5")
        ]
        
        for i, (key, formula) in enumerate(kpis, 4):
            ws[f'G{i}'] = key
            ws[f'H{i}'] = formula
            ws[f'G{i}'].style = "data"
            ws[f'H{i}'].style = "number"
        
        # Recent Activities Log
        ws['A12'] = "Recent Activities"
        ws['A12'].style = "header"
        ws.merge_cells('A12:H12')
        
        activity_headers = ["Date", "Activity Type", "Description", "Status", "User", "Duration", "Impact", "Notes"]
        for i, header in enumerate(activity_headers, 1):
            ws[f'{get_column_letter(i)}13'] = header
            ws[f'{get_column_letter(i)}13'].style = "header"
        
        # Sample activities
        sample_activities = [
            (datetime.now().strftime("%Y-%m-%d"), "Data Ingestion", "Processed new literature corpus", "Completed", "System", "15 min", "High", "Added 15 new papers"),
            ((datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d"), "Model Analysis", "Updated 3D geological interpretation", "Completed", "Analyst", "45 min", "Medium", "Revised capacity estimates"),
            ((datetime.now() - timedelta(days=2)).strftime("%Y-%m-%d"), "QA Testing", "Nightly evaluation run", "Completed", "System", "12 min", "Low", "All tests passed")
        ]
        
        for i, activity in enumerate(sample_activities, 14):
            for j, value in enumerate(activity, 1):
                ws[f'{get_column_letter(j)}{i}'] = value
                ws[f'{get_column_letter(j)}{i}'].style = "data"
        
        # Auto-fit columns (simplified to avoid merged cell issues)
        for col_num in range(1, 9):  # Columns A through H
            ws.column_dimensions[get_column_letter(col_num)].width = 20
    
    def create_literature_corpus_sheet(self):
        """Create literature corpus tracking sheet."""
        logger.info("Creating Literature Corpus sheet...")
        
        ws = self.workbook.create_sheet("Literature_Corpus")
        
        # Title and summary
        ws['A1'] = "Literature Corpus Management"
        ws['A1'].font = Font(bold=True, size=14, color=self.colors['primary'])
        ws.merge_cells('A1:I1')
        
        # Summary statistics
        ws['A2'] = "Total Documents"
        ws['B2'] = 0  # Will be updated by system
        ws['C2'] = "Last Updated"
        ws['D2'] = datetime.now().strftime("%Y-%m-%d %H:%M")
        
        ws['A3'] = "Text Chunks"
        ws['B3'] = 0  # Will be updated
        ws['C3'] = "Processing Status"
        ws['D3'] = "Active"
        
        # Headers for document tracking
        headers = [
            "Document ID", "Filename", "Title", "Authors", "Publication Date",
            "File Size (MB)", "Pages", "Processing Date", "Status", "Text Chunks",
            "Images Extracted", "Language", "Domain", "Quality Score", "Notes"
        ]
        
        for i, header in enumerate(headers, 1):
            ws[f'{get_column_letter(i)}5'] = header
            ws[f'{get_column_letter(i)}5'].style = "header"
        
        # Sample data entries
        sample_docs = [
            ("DOC001", "3D MT and Gravity Modeling - Semurup.pdf", "3D Magnetotelluric and Gravity Modeling", "Research Team", "2023-01-15", 29.2, 85, "2024-01-15", "Processed", 156, 12, "English", "Geophysics", 0.92, "High quality technical paper"),
            ("DOC002", "Geokimia - Laporan Akhir Survey Geokimia Semurup 2022_Final.pdf", "Final Geochemical Survey Report", "Field Team", "2022-12-30", 2.2, 45, "2024-01-15", "Processed", 78, 8, "Indonesian", "Geochemistry", 0.88, "Comprehensive field survey"),
            ("DOC003", "PRE FS SEMURUP (2014).pdf", "Pre-Feasibility Study Semurup", "Consulting Team", "2014-06-30", 6.2, 120, "2024-01-15", "Processed", 234, 25, "English", "Engineering", 0.85, "Economic analysis included"),
            ("DOC004", "Geothermal Energy - Inggrid Stober.pdf", "Geothermal Energy Systems", "Inggrid Stober", "2021-03-01", 24.1, 485, "2024-01-16", "Processing", 0, 0, "English", "Reference", 0.0, "Textbook reference")
        ]
        
        for i, doc in enumerate(sample_docs, 6):
            for j, value in enumerate(doc, 1):
                ws[f'{get_column_letter(j)}{i}'] = value
                ws[f'{get_column_letter(j)}{i}'].style = "data"
        
        # Status formatting
        status_col = 9  # Column I
        for row in range(6, 10):
            cell = ws[f'{get_column_letter(status_col)}{row}']
            if cell.value == "Processed":
                cell.style = "status_good"
            elif cell.value == "Processing":
                cell.style = "status_warning"
            else:
                cell.style = "status_danger"
        
        # Add data validation for status column
        status_validation = DataValidation(
            type="list",
            formula1='"Processed,Processing,Failed,Pending"',
            allow_blank=False
        )
        ws.add_data_validation(status_validation)
        status_validation.add(f'I6:I100')
        
        # Auto-fit columns (simplified to avoid merged cell issues)
        for col_num in range(1, 9):  # Columns A through H
            ws.column_dimensions[get_column_letter(col_num)].width = 20
    
    def create_image_corpus_sheet(self):
        """Create image corpus management sheet (as requested)."""
        logger.info("Creating Image Corpus sheet...")
        
        ws = self.workbook.create_sheet("Image_Corpus")
        
        # Title and summary
        ws['A1'] = "Image Corpus Management"
        ws['A1'].font = Font(bold=True, size=14, color=self.colors['primary'])
        ws.merge_cells('A1:L1')
        
        # Summary statistics
        ws['A2'] = "Total Images"
        ws['B2'] = 0  # Will be updated by system
        ws['C2'] = "Last Updated"
        ws['D2'] = datetime.now().strftime("%Y-%m-%d %H:%M")
        
        ws['A3'] = "Processed Images"
        ws['B3'] = 0  # Will be updated
        ws['C3'] = "CLIP Embeddings"
        ws['D3'] = 0  # Will be updated
        
        # Headers for image tracking
        headers = [
            "Image ID", "Filename", "Source Document", "Page Number", "Caption/Description",
            "Image Type", "Dimensions (WxH)", "File Size (KB)", "Format", "Extracted Date",
            "Processing Status", "Embedding Status", "Quality Score", "Classification", "Tags", "Notes"
        ]
        
        for i, header in enumerate(headers, 1):
            ws[f'{get_column_letter(i)}5'] = header
            ws[f'{get_column_letter(i)}5'].style = "header"
        
        # Sample image entries
        sample_images = [
            ("IMG001", "3D_MT_Gravity_p15_img1.png", "3D MT and Gravity Modeling - Semurup.pdf", 15, "Resistivity model cross-section A-A'", "Cross-section", "800x600", 245, "PNG", "2024-01-15", "Processed", "Embedded", 0.94, "Geophysics", "resistivity,cross-section,model", "High-quality technical figure"),
            ("IMG002", "3D_MT_Gravity_p23_img2.png", "3D MT and Gravity Modeling - Semurup.pdf", 23, "3D resistivity isosurfaces", "3D Visualization", "1024x768", 387, "PNG", "2024-01-15", "Processed", "Embedded", 0.91, "Geophysics", "3d,resistivity,isosurfaces", "Clear 3D representation"),
            ("IMG003", "Geokimia_p12_img1.png", "Geokimia - Laporan Akhir Survey Geokimia Semurup 2022_Final.pdf", 12, "Geochemical sampling locations map", "Map", "1200x900", 456, "PNG", "2024-01-15", "Processed", "Embedded", 0.89, "Geochemistry", "map,sampling,locations", "Detailed location map"),
            ("IMG004", "PRE_FS_p45_img3.png", "PRE FS SEMURUP (2014).pdf", 45, "Power plant layout schematic", "Schematic", "1000x700", 312, "PNG", "2024-01-15", "Processed", "Embedded", 0.86, "Engineering", "power,plant,layout", "Engineering schematic"),
            ("IMG005", "Geokimia_p8_img2.png", "Geokimia - Laporan Akhir Survey Geokimia Semurup 2022_Final.pdf", 8, "Trilinear diagram of water chemistry", "Chart", "600x600", 189, "PNG", "2024-01-15", "Processing", "Pending", 0.0, "Geochemistry", "trilinear,chemistry,water", "Geochemical analysis chart")
        ]
        
        for i, img in enumerate(sample_images, 6):
            for j, value in enumerate(img, 1):
                ws[f'{get_column_letter(j)}{i}'] = value
                ws[f'{get_column_letter(j)}{i}'].style = "data"
        
        # Status formatting
        processing_status_col = 11  # Column K
        embedding_status_col = 12   # Column L
        
        for row in range(6, 11):
            # Processing status
            proc_cell = ws[f'{get_column_letter(processing_status_col)}{row}']
            if proc_cell.value == "Processed":
                proc_cell.style = "status_good"
            elif proc_cell.value == "Processing":
                proc_cell.style = "status_warning"
            else:
                proc_cell.style = "status_danger"
            
            # Embedding status
            emb_cell = ws[f'{get_column_letter(embedding_status_col)}{row}']
            if emb_cell.value == "Embedded":
                emb_cell.style = "status_good"
            elif emb_cell.value == "Pending":
                emb_cell.style = "status_warning"
            else:
                emb_cell.style = "status_danger"
        
        # Add data validations
        proc_validation = DataValidation(
            type="list",
            formula1='"Processed,Processing,Failed,Pending"',
            allow_blank=False
        )
        ws.add_data_validation(proc_validation)
        proc_validation.add(f'K6:K100')
        
        emb_validation = DataValidation(
            type="list",
            formula1='"Embedded,Pending,Failed"',
            allow_blank=False
        )
        ws.add_data_validation(emb_validation)
        emb_validation.add(f'L6:L100')
        
        # Image type validation
        type_validation = DataValidation(
            type="list",
            formula1='"Cross-section,3D Visualization,Map,Chart,Schematic,Graph,Photo,Diagram"',
            allow_blank=False
        )
        ws.add_data_validation(type_validation)
        type_validation.add(f'F6:F100')
        
        # Auto-fit columns (simplified to avoid merged cell issues)
        for col_num in range(1, 9):  # Columns A through H
            ws.column_dimensions[get_column_letter(col_num)].width = 20
    
    def create_model_data_sheet(self):
        """Create 3D models and data inventory sheet."""
        logger.info("Creating Model Data sheet...")
        
        ws = self.workbook.create_sheet("Model_Data")
        
        # Title
        ws['A1'] = "3D Models and Data Inventory"
        ws['A1'].font = Font(bold=True, size=14, color=self.colors['primary'])
        ws.merge_cells('A1:N1')
        
        # Summary
        ws['A2'] = "Total Models"
        ws['B2'] = 0
        ws['C2'] = "Processing Status"
        ws['D2'] = "Active"
        ws['E2'] = "Coverage (%)"
        ws['F2'] = 85.2
        
        # Headers
        headers = [
            "Model ID", "Filename", "Data Type", "Property", "Grid Dimensions", 
            "Spatial Coverage", "Depth Range (m)", "File Size (MB)", "Format",
            "Processing Date", "Status", "Quality Score", "Applications", "Validation Status", "Notes"
        ]
        
        for i, header in enumerate(headers, 1):
            ws[f'{get_column_letter(i)}4'] = header
            ws[f'{get_column_letter(i)}4'].style = "header"
        
        # Sample model data
        sample_models = [
            ("MOD001", "SMP_JI03_3dmod_res_it150.dat", "Geophysical", "Resistivity", "2929x2929x41", "34.3 km²", "-4298 to 1787", 53.6, "DAT", "2024-01-15", "Processed", 0.92, "Reservoir mapping", "Validated", "High-resolution MT inversion"),
            ("MOD002", "SMP_JI03_3dmod_dens_it150.dat", "Geophysical", "Density", "2929x2929x41", "34.3 km²", "-4298 to 1787", 53.6, "DAT", "2024-01-15", "Processed", 0.88, "Structural analysis", "Validated", "Gravity inversion model"),
            ("GEO001", "semurup_geochemical_survey.xlsx", "Geochemical", "Multi-parameter", "5 locations", "Field area", "0 to -200", 0.008, "XLSX", "2024-01-15", "Processed", 0.85, "Fluid analysis", "Reviewed", "Hot spring sampling"),
            ("GEO002", "Survei Geokimia Semurup 2022.xlsx", "Geochemical", "Comprehensive", "21 samples", "Extended area", "Surface", 0.017, "XLSX", "2024-01-15", "Processed", 0.90, "Regional assessment", "Validated", "Comprehensive survey data")
        ]
        
        for i, model in enumerate(sample_models, 5):
            for j, value in enumerate(model, 1):
                ws[f'{get_column_letter(j)}{i}'] = value
                ws[f'{get_column_letter(j)}{i}'].style = "data"
        
        # Status formatting
        for row in range(5, 9):
            status_cell = ws[f'K{row}']
            if status_cell.value == "Processed":
                status_cell.style = "status_good"
            elif status_cell.value == "Processing":
                status_cell.style = "status_warning"
            else:
                status_cell.style = "status_danger"
            
            validation_cell = ws[f'N{row}']
            if validation_cell.value == "Validated":
                validation_cell.style = "status_good"
            elif validation_cell.value == "Reviewed":
                validation_cell.style = "status_warning"
        
        # Auto-fit columns (simplified to avoid merged cell issues)
        for col_num in range(1, 9):  # Columns A through H
            ws.column_dimensions[get_column_letter(col_num)].width = 20
    
    def create_engineering_results_sheet(self):
        """Create engineering analysis results sheet."""
        logger.info("Creating Engineering Results sheet...")
        
        ws = self.workbook.create_sheet("Engineering_Results")
        
        # Title
        ws['A1'] = "Engineering Analysis Results"
        ws['A1'].font = Font(bold=True, size=14, color=self.colors['primary'])
        ws.merge_cells('A1:H1')
        
        # Key Results Summary
        ws['A3'] = "Key Results Summary"
        ws['A3'].style = "header"
        ws.merge_cells('A3:B3')
        
        key_results = [
            ("Estimated Capacity (MW)", 45.2),
            ("Development Status", "Pre-commercial Development Phase"),
            ("Geothermal Potential", "Good"),
            ("Overall Confidence", "73%"),
            ("Reservoir Volume (km³)", 2.34),
            ("Caprock Thickness (m)", 125)
        ]
        
        for i, (key, value) in enumerate(key_results, 4):
            ws[f'A{i}'] = key
            ws[f'B{i}'] = value
            ws[f'A{i}'].style = "data"
            ws[f'B{i}'].style = "data"
        
        # Geological Zones Analysis
        ws['D3'] = "Geological Zones"
        ws['D3'].style = "header"
        ws.merge_cells('D3:H3')
        
        zone_headers = ["Zone Type", "Volume (km³)", "Area (km²)", "Confidence", "Description"]
        for i, header in enumerate(zone_headers):
            ws[f'{get_column_letter(i+4)}{4}'] = header
            ws[f'{get_column_letter(i+4)}{4}'].style = "header"
        
        zone_data = [
            ("Caprock", 1.12, 15.4, "85%", "Low-permeability seal"),
            ("Reservoir", 2.34, 28.7, "78%", "Primary production zone"),
            ("Basement", 4.56, 34.3, "65%", "Deep heat source"),
            ("Fracture Zones", 0.45, 8.2, "50%", "Enhanced permeability")
        ]
        
        for i, zone in enumerate(zone_data, 5):
            for j, value in enumerate(zone):
                ws[f'{get_column_letter(j+4)}{i}'] = value
                ws[f'{get_column_letter(j+4)}{i}'].style = "data"
        
        # Drilling Recommendations
        ws['A12'] = "Drilling Recommendations"
        ws['A12'].style = "header"
        ws.merge_cells('A12:H12')
        
        drill_headers = ["Priority", "Target Type", "Location", "Depth Range (m)", "Estimated Capacity (MW)", "Confidence", "Rationale"]
        for i, header in enumerate(drill_headers, 1):
            ws[f'{get_column_letter(i)}{13}'] = header
            ws[f'{get_column_letter(i)}{13}'].style = "header"
        
        drill_recs = [
            (1, "Production Well", "Reservoir Zone A", "1500-2500", 15.0, "80%", "Primary reservoir target"),
            (2, "Enhanced Production", "Fracture Intersection", "1200-2200", 18.0, "70%", "Natural fracture enhancement"),
            (3, "Injection Well", "Caprock Margin", "500-1500", 0, "85%", "Sustainable operation"),
            (4, "Exploration Well", "Deep Basement", "2500-3500", 0, "60%", "Heat source exploration")
        ]
        
        for i, rec in enumerate(drill_recs, 14):
            for j, value in enumerate(rec, 1):
                ws[f'{get_column_letter(j)}{i}'] = value
                ws[f'{get_column_letter(j)}{i}'].style = "data"
        
        # Auto-fit columns (simplified to avoid merged cell issues)
        for col_num in range(1, 9):  # Columns A through H
            ws.column_dimensions[get_column_letter(col_num)].width = 20
    
    def create_qa_performance_sheet(self):
        """Create QA system performance tracking sheet."""
        logger.info("Creating QA Performance sheet...")
        
        ws = self.workbook.create_sheet("QA_Performance")
        
        # Title
        ws['A1'] = "QA System Performance Metrics"
        ws['A1'].font = Font(bold=True, size=14, color=self.colors['primary'])
        ws.merge_cells('A1:J1')
        
        # Performance Summary
        ws['A2'] = "Total Questions"
        ws['B2'] = 0
        ws['C2'] = "Success Rate"
        ws['D2'] = "95.2%"
        ws['E2'] = "Avg Response Time"
        ws['F2'] = "3.2s"
        
        # Performance metrics over time
        ws['A4'] = "Performance History"
        ws['A4'].style = "header"
        ws.merge_cells('A4:J4')
        
        perf_headers = ["Date", "Questions Tested", "Success Rate (%)", "Avg Response Time (s)", 
                       "Avg Confidence (%)", "Citations per Answer", "Context Sources Used", 
                       "Literature Coverage (%)", "Model Data Usage (%)", "Notes"]
        
        for i, header in enumerate(perf_headers, 1):
            ws[f'{get_column_letter(i)}5'] = header
            ws[f'{get_column_letter(i)}5'].style = "header"
        
        # Sample performance data
        perf_data = [
            ("2024-01-15", 25, 96.0, 3.1, 78.5, 3.2, 4.1, 85.0, 72.0, "All tests passed"),
            ("2024-01-14", 22, 95.5, 3.4, 76.8, 3.0, 3.8, 83.0, 68.0, "Minor timeout issues"),
            ("2024-01-13", 28, 92.9, 3.8, 74.2, 2.9, 3.5, 80.0, 65.0, "Performance regression"),
            ("2024-01-12", 26, 94.2, 3.2, 77.1, 3.1, 3.9, 82.0, 70.0, "System optimized"),
            ("2024-01-11", 24, 93.8, 3.6, 75.5, 2.8, 3.6, 78.0, 63.0, "New model integrated")
        ]
        
        for i, data in enumerate(perf_data, 6):
            for j, value in enumerate(data, 1):
                ws[f'{get_column_letter(j)}{i}'] = value
                ws[f'{get_column_letter(j)}{i}'].style = "data"
        
        # Question Category Performance
        ws['A13'] = "Category Performance Analysis"
        ws['A13'].style = "header"
        ws.merge_cells('A13:G13')
        
        cat_headers = ["Category", "Questions", "Success Rate (%)", "Avg Confidence (%)", 
                      "Avg Response Time (s)", "Primary Sources", "Quality Score"]
        
        for i, header in enumerate(cat_headers, 1):
            ws[f'{get_column_letter(i)}14'] = header
            ws[f'{get_column_letter(i)}14'].style = "header"
        
        category_data = [
            ("Capacity Assessment", 12, 98.0, 85.2, 4.1, "Engineering Analysis", 0.92),
            ("Geological Analysis", 15, 94.5, 82.1, 5.2, "3D Models + Literature", 0.89),
            ("Drilling Recommendations", 8, 96.0, 78.9, 3.8, "Engineering + Models", 0.87),
            ("Geochemistry", 6, 92.0, 76.4, 2.9, "Literature + Data", 0.85),
            ("Complex Integration", 4, 88.0, 74.2, 8.1, "All Sources", 0.82)
        ]
        
        for i, cat in enumerate(category_data, 15):
            for j, value in enumerate(cat, 1):
                ws[f'{get_column_letter(j)}{i}'] = value
                ws[f'{get_column_letter(j)}{i}'].style = "data"
        
        # Auto-fit columns (simplified to avoid merged cell issues)
        for col_num in range(1, 9):  # Columns A through H
            ws.column_dimensions[get_column_letter(col_num)].width = 20
    
    def create_evaluation_results_sheet(self):
        """Create evaluation results tracking sheet."""
        logger.info("Creating Evaluation Results sheet...")
        
        ws = self.workbook.create_sheet("Evaluation_Results")
        
        # Title
        ws['A1'] = "System Evaluation Results"
        ws['A1'].font = Font(bold=True, size=14, color=self.colors['primary'])
        ws.merge_cells('A1:L1')
        
        # Latest Evaluation Summary
        ws['A2'] = "System Health Score"
        ws['B2'] = "92.5%"
        ws['C2'] = "Last Evaluation"
        ws['D2'] = datetime.now().strftime("%Y-%m-%d %H:%M")
        ws['E2'] = "Status"
        ws['F2'] = "Healthy"
        
        # Evaluation History
        ws['A4'] = "Evaluation History"
        ws['A4'].style = "header"
        ws.merge_cells('A4:L4')
        
        eval_headers = ["Evaluation ID", "Date", "Duration (min)", "Overall Status", "Pass Rate (%)", 
                       "Smoke Tests", "Integration Tests", "Performance Tests", "Quality Tests", 
                       "Critical Issues", "Recommendations", "Notes"]
        
        for i, header in enumerate(eval_headers, 1):
            ws[f'{get_column_letter(i)}5'] = header
            ws[f'{get_column_letter(i)}5'].style = "header"
        
        # Sample evaluation data
        eval_data = [
            ("eval_20240115_020000", "2024-01-15 02:00", 28.5, "PASSED", 95.2, "PASS", "PASS", "PASS", "PASS", 0, "Monitor performance", "All systems nominal"),
            ("eval_20240114_020000", "2024-01-14 02:00", 32.1, "PASSED", 92.8, "PASS", "PASS", "DEGRADED", "PASS", 1, "Review response times", "Minor performance issues"),
            ("eval_20240113_020000", "2024-01-13 02:00", 25.8, "FAILED", 87.5, "PASS", "FAIL", "FAIL", "PASS", 3, "System maintenance required", "Multiple component failures"),
            ("eval_20240112_020000", "2024-01-12 02:00", 29.3, "PASSED", 94.1, "PASS", "PASS", "PASS", "PASS", 0, "Continue monitoring", "Good system performance")
        ]
        
        for i, eval_result in enumerate(eval_data, 6):
            for j, value in enumerate(eval_result, 1):
                ws[f'{get_column_letter(j)}{i}'] = value
                ws[f'{get_column_letter(j)}{i}'].style = "data"
        
        # Status formatting
        status_columns = [4, 6, 7, 8, 9]  # Overall, Smoke, Integration, Performance, Quality
        for row in range(6, 10):
            for col in status_columns:
                cell = ws[f'{get_column_letter(col)}{row}']
                if cell.value in ["PASSED", "PASS"]:
                    cell.style = "status_good"
                elif cell.value in ["DEGRADED", "WARNING"]:
                    cell.style = "status_warning"
                else:
                    cell.style = "status_danger"
        
        # System Health Metrics
        ws['A12'] = "System Health Metrics"
        ws['A12'].style = "header"
        ws.merge_cells('A12:F12')
        
        health_metrics = [
            ("CPU Usage (%)", 15.2),
            ("Memory Usage (%)", 68.4),
            ("Disk Usage (%)", 45.7),
            ("Response Time (ms)", 3200),
            ("Uptime (hours)", 720),
            ("Error Rate (%)", 2.1)
        ]
        
        for i, (metric, value) in enumerate(health_metrics, 13):
            ws[f'A{i}'] = metric
            ws[f'B{i}'] = value
            ws[f'A{i}'].style = "data"
            ws[f'B{i}'].style = "number"
        
        # Auto-fit columns (simplified to avoid merged cell issues)
        for col_num in range(1, 9):  # Columns A through H
            ws.column_dimensions[get_column_letter(col_num)].width = 20
    
    def create_risk_assessment_sheet(self):
        """Create risk assessment and management sheet."""
        logger.info("Creating Risk Assessment sheet...")
        
        ws = self.workbook.create_sheet("Risk_Assessment")
        
        # Title
        ws['A1'] = "Risk Assessment and Management"
        ws['A1'].font = Font(bold=True, size=14, color=self.colors['primary'])
        ws.merge_cells('A1:J1')
        
        # Risk Summary
        ws['A3'] = "Risk Summary"
        ws['A3'].style = "header"
        ws.merge_cells('A3:D3')
        
        ws['A4'] = "Total Risks Identified"
        ws['B4'] = 12
        ws['C4'] = "High Priority Risks"
        ws['D4'] = 3
        
        ws['A5'] = "Risk Score (Overall)"
        ws['B5'] = "Medium"
        ws['C5'] = "Last Assessment"
        ws['D5'] = datetime.now().strftime("%Y-%m-%d")
        
        # Risk Register
        ws['A7'] = "Risk Register"
        ws['A7'].style = "header"
        ws.merge_cells('A7:J7')
        
        risk_headers = ["Risk ID", "Category", "Description", "Probability", "Impact", 
                       "Risk Score", "Mitigation Strategy", "Owner", "Status", "Review Date"]
        
        for i, header in enumerate(risk_headers, 1):
            ws[f'{get_column_letter(i)}8'] = header
            ws[f'{get_column_letter(i)}8'].style = "header"
        
        # Sample risk data
        risk_data = [
            ("R001", "Technical", "API rate limiting affecting performance", "Medium", "High", "High", "Implement caching and request optimization", "Tech Team", "Active", "2024-02-15"),
            ("R002", "Data", "Literature corpus incompleteness", "High", "Medium", "High", "Continuous literature monitoring and acquisition", "Data Team", "Active", "2024-01-30"),
            ("R003", "System", "Hardware failure affecting availability", "Low", "High", "Medium", "Redundancy and backup systems implementation", "Ops Team", "Active", "2024-03-01"),
            ("R004", "Quality", "Model interpretation accuracy degradation", "Medium", "Medium", "Medium", "Regular validation and expert review", "Analysis Team", "Monitor", "2024-02-01"),
            ("R005", "Security", "Unauthorized access to sensitive data", "Low", "High", "Medium", "Access controls and audit logging", "Security Team", "Active", "2024-01-20"),
            ("R006", "Operational", "Key personnel availability", "Medium", "Medium", "Medium", "Knowledge documentation and cross-training", "HR Team", "Monitor", "2024-02-10")
        ]
        
        for i, risk in enumerate(risk_data, 9):
            for j, value in enumerate(risk, 1):
                ws[f'{get_column_letter(j)}{i}'] = value
                ws[f'{get_column_letter(j)}{i}'].style = "data"
        
        # Risk score formatting
        risk_score_col = 6
        for row in range(9, 15):
            cell = ws[f'{get_column_letter(risk_score_col)}{row}']
            if cell.value == "High":
                cell.style = "status_danger"
            elif cell.value == "Medium":
                cell.style = "status_warning"
            else:
                cell.style = "status_good"
        
        # Add data validations
        prob_validation = DataValidation(
            type="list",
            formula1='"Low,Medium,High"',
            allow_blank=False
        )
        ws.add_data_validation(prob_validation)
        prob_validation.add('D9:D100')
        
        impact_validation = DataValidation(
            type="list",
            formula1='"Low,Medium,High"',
            allow_blank=False
        )
        ws.add_data_validation(impact_validation)
        impact_validation.add('E9:E100')
        
        status_validation = DataValidation(
            type="list",
            formula1='"Active,Monitor,Closed,Escalated"',
            allow_blank=False
        )
        ws.add_data_validation(status_validation)
        status_validation.add('I9:I100')
        
        # Auto-fit columns (simplified to avoid merged cell issues)
        for col_num in range(1, 9):  # Columns A through H
            ws.column_dimensions[get_column_letter(col_num)].width = 20
    
    def create_configuration_sheet(self):
        """Create system configuration management sheet."""
        logger.info("Creating Configuration Management sheet...")
        
        ws = self.workbook.create_sheet("Configuration")
        
        # Title
        ws['A1'] = "System Configuration Management"
        ws['A1'].font = Font(bold=True, size=14, color=self.colors['primary'])
        ws.merge_cells('A1:H1')
        
        # System Configuration
        ws['A3'] = "Current System Configuration"
        ws['A3'].style = "header"
        ws.merge_cells('A3:D3')
        
        config_data = [
            ("System Version", "v2.0.0"),
            ("Python Version", "3.13.5"),
            ("OpenAI Model", "GPT-o3"),
            ("Embedding Model", "text-embedding-ada-002"),
            ("CLIP Model", "ViT-B-32"),
            ("Vector Database", "ChromaDB v0.4.15"),
            ("3D Processing", "PyVista v0.42.0"),
            ("Web Framework", "FastAPI v0.100.0"),
            ("Last Updated", datetime.now().strftime("%Y-%m-%d %H:%M"))
        ]
        
        for i, (key, value) in enumerate(config_data, 4):
            ws[f'A{i}'] = key
            ws[f'B{i}'] = value
            ws[f'A{i}'].style = "data"
            ws[f'B{i}'].style = "data"
        
        # Environment Variables
        ws['D3'] = "Environment Configuration"
        ws['D3'].style = "header"
        ws.merge_cells('D3:F3')
        
        env_data = [
            ("OPENAI_API_KEY", "Set (Hidden)"),
            ("MAX_CHUNK_SIZE", "1000"),
            ("CHUNK_OVERLAP", "200"),
            ("TOP_K_TEXTS", "5"),
            ("TOP_K_IMAGES", "3"),
            ("SIMILARITY_THRESHOLD", "0.7"),
            ("VTU_COMPRESSION", "True"),
            ("DEFAULT_TEMPERATURE", "0.3")
        ]
        
        for i, (key, value) in enumerate(env_data, 4):
            ws[f'D{i}'] = key
            ws[f'E{i}'] = value
            ws[f'D{i}'].style = "data"
            ws[f'E{i}'].style = "data"
        
        # Configuration Change Log
        ws['A15'] = "Configuration Change Log"
        ws['A15'].style = "header"
        ws.merge_cells('A15:H15')
        
        change_headers = ["Date", "Component", "Parameter", "Old Value", "New Value", "Reason", "Changed By", "Approved By"]
        
        for i, header in enumerate(change_headers, 1):
            ws[f'{get_column_letter(i)}16'] = header
            ws[f'{get_column_letter(i)}16'].style = "header"
        
        # Sample change log
        change_data = [
            ("2024-01-15", "QA Server", "Temperature", "0.5", "0.3", "Improve response consistency", "Admin", "Tech Lead"),
            ("2024-01-14", "Text Processing", "Chunk Size", "800", "1000", "Better context coverage", "Admin", "Data Lead"),
            ("2024-01-13", "Embeddings", "Model", "ada-001", "ada-002", "Performance improvement", "Admin", "ML Lead"),
            ("2024-01-12", "Question Graph", "TOP_K_SIMILAR", "5", "10", "Improved question routing", "Admin", "Tech Lead")
        ]
        
        for i, change in enumerate(change_data, 17):
            for j, value in enumerate(change, 1):
                ws[f'{get_column_letter(j)}{i}'] = value
                ws[f'{get_column_letter(j)}{i}'].style = "data"
        
        # Auto-fit columns (simplified to avoid merged cell issues)
        for col_num in range(1, 9):  # Columns A through H
            ws.column_dimensions[get_column_letter(col_num)].width = 20
    
    def generate_complete_template(self, output_file: Optional[Path] = None) -> Path:
        """Generate the complete Excel template."""
        logger.info("=== Generating Complete Excel Template ===")
        
        # Create all sheets
        self.create_project_overview_sheet()
        self.create_literature_corpus_sheet()
        self.create_image_corpus_sheet()  # As requested
        self.create_model_data_sheet()
        self.create_engineering_results_sheet()
        self.create_qa_performance_sheet()
        self.create_evaluation_results_sheet()
        self.create_risk_assessment_sheet()
        self.create_configuration_sheet()
        
        # Save the workbook
        output_path = output_file or TEMPLATE_FILE
        
        try:
            self.workbook.save(output_path)
            logger.info(f"Excel template saved to: {output_path}")
            
            # Report statistics
            sheet_count = len(self.workbook.sheetnames)
            file_size = output_path.stat().st_size / 1024  # KB
            
            logger.info(f"Template created with {sheet_count} sheets ({file_size:.1f} KB)")
            
            return output_path
            
        except Exception as e:
            logger.error(f"Failed to save Excel template: {e}")
            raise
    
    def update_template_with_system_data(self, template_path: Path):
        """Update template with actual system data."""
        logger.info("Updating template with system data...")
        
        try:
            # Load existing data if available
            data_sources = {
                'text_collection_count': 0,
                'image_collection_count': 0,
                'models_count': 0,
                'summaries_available': False
            }
            
            # Try to get actual counts
            try:
                import chromadb
                
                # Text collection
                if Path("knowledge/text_emb").exists():
                    text_client = chromadb.PersistentClient(path="knowledge/text_emb")
                    try:
                        text_collection = text_client.get_collection("geothermal_texts")
                        data_sources['text_collection_count'] = text_collection.count()
                    except:
                        pass
                
                # Image collection
                if Path("knowledge/image_emb").exists():
                    image_client = chromadb.PersistentClient(path="knowledge/image_emb")
                    try:
                        image_collection = image_client.get_collection("geothermal_images")
                        data_sources['image_collection_count'] = image_collection.count()
                    except:
                        pass
                
                # Models count
                models_dir = Path("digital_twin/grids")
                if models_dir.exists():
                    data_sources['models_count'] = len(list(models_dir.glob("*.vtu")))
                
                # Summaries
                summaries_file = Path("digital_twin/cache/twin_summaries.yaml")
                data_sources['summaries_available'] = summaries_file.exists()
                
            except ImportError:
                logger.warning("ChromaDB not available for data update")
            
            # Update template with actual data
            wb = openpyxl.load_workbook(template_path)
            
            # Update Literature Corpus sheet
            if "Literature_Corpus" in wb.sheetnames:
                ws = wb["Literature_Corpus"]
                ws['B2'] = data_sources['text_collection_count']
            
            # Update Image Corpus sheet
            if "Image_Corpus" in wb.sheetnames:
                ws = wb["Image_Corpus"]
                ws['B2'] = data_sources['image_collection_count']
                ws['D3'] = data_sources['image_collection_count']
            
            # Update Model Data sheet
            if "Model_Data" in wb.sheetnames:
                ws = wb["Model_Data"]
                ws['B2'] = data_sources['models_count']
            
            # Save updated template
            wb.save(template_path)
            logger.info("Template updated with system data")
            
        except Exception as e:
            logger.error(f"Failed to update template with system data: {e}")


def main():
    """Main entry point."""
    import argparse
    
    parser = argparse.ArgumentParser(description="Excel Template Generator")
    parser.add_argument("--output", type=str, help="Output file path")
    parser.add_argument("--update-existing", action="store_true", help="Update existing template with system data")
    
    args = parser.parse_args()
    
    try:
        generator = GeothermalExcelGenerator()
        
        if args.update_existing:
            # Update existing template
            template_path = Path(args.output) if args.output else TEMPLATE_FILE
            if template_path.exists():
                generator.update_template_with_system_data(template_path)
                print(f"Template updated: {template_path}")
            else:
                print(f"Template not found: {template_path}")
        else:
            # Generate new template
            output_file = Path(args.output) if args.output else None
            template_path = generator.generate_complete_template(output_file)
            
            # Update with system data
            generator.update_template_with_system_data(template_path)
            
            print(f"\n=== Excel Template Generated ===")
            print(f"File: {template_path}")
            print(f"Sheets: {len(generator.workbook.sheetnames)}")
            print(f"Size: {template_path.stat().st_size / 1024:.1f} KB")
            
            print(f"\nSheet Summary:")
            for sheet_name in generator.workbook.sheetnames:
                print(f"  • {sheet_name}")
            
            print(f"\nKey Features:")
            print(f"  • Project overview dashboard")
            print(f"  • Literature and image corpus tracking")
            print(f"  • 3D model inventory management")
            print(f"  • Engineering analysis results")
            print(f"  • QA performance monitoring")
            print(f"  • System evaluation tracking")
            print(f"  • Risk assessment matrix")
            print(f"  • Configuration management")
            
    except KeyboardInterrupt:
        logger.info("Template generation interrupted")
        sys.exit(0)
    except Exception as e:
        logger.error(f"Fatal error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()