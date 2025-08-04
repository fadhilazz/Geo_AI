#!/usr/bin/env python3
"""
Simple Excel Template Generator - Final MAX Version

Creates the comprehensive Excel template for geothermal digital twin governance.
"""

import logging
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def create_excel_template():
    """Create comprehensive Excel template."""
    logger.info("Creating comprehensive Excel template...")
    
    # Ensure directory exists
    excel_dir = Path("excel")
    excel_dir.mkdir(exist_ok=True)
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Color scheme
    colors = {
        'header': 'FF17A2B8',
        'primary': 'FF2E4BC4',
        'success': 'FF198754',
        'warning': 'FFFFC107',
        'danger': 'FFDC3545'
    }
    
    # 1. Project Overview Sheet
    ws1 = wb.create_sheet("Project_Overview")
    
    # Title
    ws1['A1'] = "Geothermal Digital Twin AI - Project Dashboard"
    ws1['A1'].font = Font(bold=True, size=16, color=colors['primary'])
    
    # Project Information
    ws1['A3'] = "Project Information"
    ws1['A3'].font = Font(bold=True, size=12)
    ws1['A3'].fill = PatternFill(start_color=colors['header'], end_color=colors['header'], fill_type='solid')
    
    project_info = [
        ("Project Name", "Semurup Geothermal Digital Twin"),
        ("Location", "Semurup, Jambi, Sumatra, Indonesia"),
        ("Project Phase", "Development"),
        ("Last Updated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("System Version", "v2.0 MAX"),
        ("Status", "Active")
    ]
    
    for i, (key, value) in enumerate(project_info, 4):
        ws1[f'A{i}'] = key
        ws1[f'B{i}'] = value
    
    # 2. Literature Corpus Sheet
    ws2 = wb.create_sheet("Literature_Corpus")
    
    ws2['A1'] = "Literature Corpus Management"
    ws2['A1'].font = Font(bold=True, size=14, color=colors['primary'])
    
    # Headers
    lit_headers = ["Document ID", "Filename", "Title", "Authors", "Publication Date", 
                   "Pages", "Processing Date", "Status", "Text Chunks", "Images Extracted"]
    
    for i, header in enumerate(lit_headers, 1):
        cell = ws2[f'{get_column_letter(i)}3']
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=colors['header'], end_color=colors['header'], fill_type='solid')
    
    # Sample data
    sample_docs = [
        ("DOC001", "3D MT and Gravity Modeling - Semurup.pdf", "3D Magnetotelluric Analysis", "Research Team", "2023-01-15", 85, "2024-01-15", "Processed", 156, 12),
        ("DOC002", "Geokimia - Laporan Akhir Survey.pdf", "Geochemical Survey Report", "Field Team", "2022-12-30", 45, "2024-01-15", "Processed", 78, 8),
        ("DOC003", "PRE FS SEMURUP (2014).pdf", "Pre-Feasibility Study", "Consulting Team", "2014-06-30", 120, "2024-01-15", "Processed", 234, 25)
    ]
    
    for i, doc in enumerate(sample_docs, 4):
        for j, value in enumerate(doc, 1):
            ws2[f'{get_column_letter(j)}{i}'] = value
    
    # 3. Image Corpus Sheet (as requested)
    ws3 = wb.create_sheet("Image_Corpus")
    
    ws3['A1'] = "Image Corpus Management"
    ws3['A1'].font = Font(bold=True, size=14, color=colors['primary'])
    
    # Image tracking headers
    img_headers = ["Image ID", "Filename", "Source Document", "Page", "Caption", "Type", 
                   "Dimensions", "Size (KB)", "Status", "Embedding Status", "Quality Score"]
    
    for i, header in enumerate(img_headers, 1):
        cell = ws3[f'{get_column_letter(i)}3']
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=colors['header'], end_color=colors['header'], fill_type='solid')
    
    # Sample image data
    sample_images = [
        ("IMG001", "3D_MT_Gravity_p15_img1.png", "3D MT and Gravity Modeling.pdf", 15, "Resistivity cross-section", "Cross-section", "800x600", 245, "Processed", "Embedded", 0.94),
        ("IMG002", "Geokimia_p12_img1.png", "Geochemical Survey Report.pdf", 12, "Sampling locations map", "Map", "1200x900", 456, "Processed", "Embedded", 0.89),
        ("IMG003", "PRE_FS_p45_img3.png", "Pre-Feasibility Study.pdf", 45, "Power plant layout", "Schematic", "1000x700", 312, "Processed", "Embedded", 0.86)
    ]
    
    for i, img in enumerate(sample_images, 4):
        for j, value in enumerate(img, 1):
            ws3[f'{get_column_letter(j)}{i}'] = value
    
    # 4. Model Data Sheet
    ws4 = wb.create_sheet("Model_Data")
    
    ws4['A1'] = "3D Models and Data Inventory"
    ws4['A1'].font = Font(bold=True, size=14, color=colors['primary'])
    
    model_headers = ["Model ID", "Filename", "Data Type", "Property", "Grid Dimensions", 
                     "Coverage", "Depth Range", "Size (MB)", "Status", "Quality Score"]
    
    for i, header in enumerate(model_headers, 1):
        cell = ws4[f'{get_column_letter(i)}3']
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=colors['header'], end_color=colors['header'], fill_type='solid')
    
    # Sample model data
    sample_models = [
        ("MOD001", "SMP_JI03_3dmod_res_it150.dat", "Geophysical", "Resistivity", "2929x2929x41", "34.3 km²", "-4298 to 1787m", 53.6, "Processed", 0.92),
        ("MOD002", "SMP_JI03_3dmod_dens_it150.dat", "Geophysical", "Density", "2929x2929x41", "34.3 km²", "-4298 to 1787m", 53.6, "Processed", 0.88),
        ("GEO001", "semurup_geochemical_survey.xlsx", "Geochemical", "Multi-parameter", "5 locations", "Field area", "0 to -200m", 0.008, "Processed", 0.85)
    ]
    
    for i, model in enumerate(sample_models, 4):
        for j, value in enumerate(model, 1):
            ws4[f'{get_column_letter(j)}{i}'] = value
    
    # 5. Engineering Results Sheet
    ws5 = wb.create_sheet("Engineering_Results")
    
    ws5['A1'] = "Engineering Analysis Results"
    ws5['A1'].font = Font(bold=True, size=14, color=colors['primary'])
    
    # Key results
    key_results = [
        ("Estimated Capacity (MW)", 45.2),
        ("Development Status", "Pre-commercial Development Phase"),
        ("Geothermal Potential", "Good"),
        ("Overall Confidence", "73%"),
        ("Reservoir Volume (km³)", 2.34),
        ("Caprock Thickness (m)", 125)
    ]
    
    ws5['A3'] = "Key Results Summary"
    ws5['A3'].font = Font(bold=True)
    
    for i, (key, value) in enumerate(key_results, 4):
        ws5[f'A{i}'] = key
        ws5[f'B{i}'] = value
    
    # 6. QA Performance Sheet
    ws6 = wb.create_sheet("QA_Performance")
    
    ws6['A1'] = "QA System Performance Metrics"
    ws6['A1'].font = Font(bold=True, size=14, color=colors['primary'])
    
    perf_headers = ["Date", "Questions Tested", "Success Rate (%)", "Avg Response Time (s)", 
                   "Avg Confidence (%)", "Citations per Answer", "Notes"]
    
    ws6['A3'] = "Performance History"
    ws6['A3'].font = Font(bold=True)
    
    for i, header in enumerate(perf_headers, 1):
        cell = ws6[f'{get_column_letter(i)}4']
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=colors['header'], end_color=colors['header'], fill_type='solid')
    
    # Sample performance data
    perf_data = [
        ("2024-01-15", 25, 96.0, 3.1, 78.5, 3.2, "All tests passed"),
        ("2024-01-14", 22, 95.5, 3.4, 76.8, 3.0, "Minor timeout issues"),
        ("2024-01-13", 28, 92.9, 3.8, 74.2, 2.9, "Performance regression")
    ]
    
    for i, data in enumerate(perf_data, 5):
        for j, value in enumerate(data, 1):
            ws6[f'{get_column_letter(j)}{i}'] = value
    
    # 7. Evaluation Results Sheet
    ws7 = wb.create_sheet("Evaluation_Results")
    
    ws7['A1'] = "System Evaluation Results"
    ws7['A1'].font = Font(bold=True, size=14, color=colors['primary'])
    
    eval_headers = ["Evaluation ID", "Date", "Duration (min)", "Overall Status", "Pass Rate (%)", 
                   "Critical Issues", "Notes"]
    
    ws7['A3'] = "Evaluation History"
    ws7['A3'].font = Font(bold=True)
    
    for i, header in enumerate(eval_headers, 1):
        cell = ws7[f'{get_column_letter(i)}4']
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=colors['header'], end_color=colors['header'], fill_type='solid')
    
    # Sample evaluation data
    eval_data = [
        ("eval_20240115_020000", "2024-01-15 02:00", 28.5, "PASSED", 95.2, 0, "All systems nominal"),
        ("eval_20240114_020000", "2024-01-14 02:00", 32.1, "PASSED", 92.8, 1, "Minor performance issues"),
        ("eval_20240113_020000", "2024-01-13 02:00", 25.8, "FAILED", 87.5, 3, "Multiple component failures")
    ]
    
    for i, eval_result in enumerate(eval_data, 5):
        for j, value in enumerate(eval_result, 1):
            ws7[f'{get_column_letter(j)}{i}'] = value
    
    # 8. Risk Assessment Sheet
    ws8 = wb.create_sheet("Risk_Assessment")
    
    ws8['A1'] = "Risk Assessment and Management"
    ws8['A1'].font = Font(bold=True, size=14, color=colors['primary'])
    
    risk_headers = ["Risk ID", "Category", "Description", "Probability", "Impact", "Risk Score", "Mitigation Strategy", "Status"]
    
    ws8['A3'] = "Risk Register"
    ws8['A3'].font = Font(bold=True)
    
    for i, header in enumerate(risk_headers, 1):
        cell = ws8[f'{get_column_letter(i)}4']
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=colors['header'], end_color=colors['header'], fill_type='solid')
    
    # Sample risk data
    risk_data = [
        ("R001", "Technical", "API rate limiting affecting performance", "Medium", "High", "High", "Implement caching and optimization", "Active"),
        ("R002", "Data", "Literature corpus incompleteness", "High", "Medium", "High", "Continuous monitoring and acquisition", "Active"),
        ("R003", "System", "Hardware failure affecting availability", "Low", "High", "Medium", "Redundancy and backup systems", "Active")
    ]
    
    for i, risk in enumerate(risk_data, 5):
        for j, value in enumerate(risk, 1):
            ws8[f'{get_column_letter(j)}{i}'] = value
    
    # 9. Configuration Sheet
    ws9 = wb.create_sheet("Configuration")
    
    ws9['A1'] = "System Configuration Management"
    ws9['A1'].font = Font(bold=True, size=14, color=colors['primary'])
    
    # Configuration data
    config_data = [
        ("System Version", "v2.0 MAX"),
        ("Python Version", "3.13.5"),
        ("OpenAI Model", "GPT-o3"),
        ("Embedding Model", "text-embedding-ada-002"),
        ("CLIP Model", "ViT-B-32"),
        ("Vector Database", "ChromaDB v0.4.15"),
        ("3D Processing", "PyVista v0.42.0"),
        ("Web Framework", "FastAPI v0.100.0"),
        ("Last Updated", datetime.now().strftime("%Y-%m-%d %H:%M"))
    ]
    
    ws9['A3'] = "Current System Configuration"
    ws9['A3'].font = Font(bold=True)
    
    for i, (key, value) in enumerate(config_data, 4):
        ws9[f'A{i}'] = key
        ws9[f'B{i}'] = value
    
    # Set column widths for all sheets
    for ws in wb.worksheets:
        for col_num in range(1, 12):  # Columns A through K
            ws.column_dimensions[get_column_letter(col_num)].width = 20
    
    # Save the workbook
    output_file = Path("excel/geothermal_AI_template_v2.xlsx")
    wb.save(output_file)
    
    logger.info(f"Excel template created: {output_file}")
    logger.info(f"Sheets: {len(wb.worksheets)}")
    logger.info(f"Size: {output_file.stat().st_size / 1024:.1f} KB")
    
    return output_file

if __name__ == "__main__":
    try:
        template_path = create_excel_template()
        
        print(f"\n🎉 COMPREHENSIVE EXCEL TEMPLATE CREATED!")
        print(f"📁 File: {template_path}")
        print(f"📊 Sheets: 9 comprehensive governance sheets")
        print(f"💾 Size: {template_path.stat().st_size / 1024:.1f} KB")
        
        print(f"\n📋 Sheet Overview:")
        print(f"  1. Project_Overview - Dashboard and key metrics")
        print(f"  2. Literature_Corpus - Document tracking and management")
        print(f"  3. Image_Corpus - Image processing and embedding status")
        print(f"  4. Model_Data - 3D models and data inventory")
        print(f"  5. Engineering_Results - Analysis results and recommendations")
        print(f"  6. QA_Performance - System performance metrics")
        print(f"  7. Evaluation_Results - Automated evaluation tracking")
        print(f"  8. Risk_Assessment - Risk management matrix")
        print(f"  9. Configuration - System configuration management")
        
        print(f"\n✅ READY FOR PRODUCTION USE!")
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()